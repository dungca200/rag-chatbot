import logging
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class XLSXParser:
    """Parser for Excel spreadsheets using openpyxl."""

    SUPPORTED_EXTENSIONS = ['.xlsx', '.xlsm']

    def parse(self, file_path: str) -> Dict:
        """
        Extract content from an XLSX file.

        Args:
            file_path: Path to the XLSX file

        Returns:
            Dict with content, metadata, and status
        """
        path = Path(file_path)

        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {file_path}",
                "content": "",
                "metadata": {}
            }

        if path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
            return {
                "success": False,
                "error": f"Unsupported file type: {path.suffix}",
                "content": "",
                "metadata": {}
            }

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets_content = []
            total_rows = 0

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []

                for row in sheet.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    if any(cell is not None for cell in row):
                        row_text = " | ".join(
                            str(cell) if cell is not None else ""
                            for cell in row
                        )
                        rows.append(row_text)
                        total_rows += 1

                if rows:
                    sheet_content = f"[Sheet: {sheet_name}]\n" + "\n".join(rows)
                    sheets_content.append(sheet_content)

            wb.close()

            content = "\n\n".join(sheets_content)

            metadata = {
                "filename": path.name,
                "file_type": "xlsx",
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "total_rows": total_rows,
                "file_size": path.stat().st_size
            }

            logger.info(f"Parsed XLSX: {path.name}, {len(wb.sheetnames)} sheets, {total_rows} rows")

            return {
                "success": True,
                "content": content,
                "metadata": metadata,
                "error": None
            }

        except Exception as e:
            logger.error(f"Error parsing XLSX {file_path}: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "content": "",
                "metadata": {"filename": path.name}
            }


def parse_xlsx(file_path: str) -> Dict:
    """Convenience function for XLSX parsing."""
    parser = XLSXParser()
    return parser.parse(file_path)
